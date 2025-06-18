from django.contrib import admin
from django.contrib import admin
from django import forms
from django.contrib.admin import AdminSite
from .models import MealRecord, Student, ClassRoom, StudentPayment,MealPrice,AuditLog
from django.contrib.auth.models import User, Group
from django.contrib.auth.admin import UserAdmin, GroupAdmin
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path
from django.contrib import messages
from openpyxl import Workbook, load_workbook
import csv
from django.urls import reverse
import json
import openpyxl
from django.shortcuts import redirect,render
from django.shortcuts import render
from datetime import date, timedelta,datetime
from decimal import Decimal
from django.db.models import Q
from django_admin_listfilter_dropdown.filters import (
    DropdownFilter, RelatedDropdownFilter)
admin.site.site_header  = "Trang quản trị bữa ăn học sinh"
admin.site.site_title   = "Quản lý bữa ăn"
admin.site.index_title  = "Bảng điều khiển"
class TermFilter(admin.SimpleListFilter):
    title            = 'Học kỳ/Niên khoá'
    parameter_name   = 'term'

    def lookups(self, request, model_admin):
        # Lấy tất cả các term, order_by('-term') để term mới nhất lên đầu
        terms = (
            model_admin.model.objects
            .values_list('term', flat=True)
            .distinct()
            .order_by('-term')
        )
        return [(t, t) for t in terms]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(term=self.value())
        return queryset
class AuditLogMixin:
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        AuditLog.objects.create(
            user=request.user,
            action='change' if change else 'add',
            path=request.path,
            data=json.dumps({
                'model': obj._meta.model_name,
                'object_id': obj.pk,
                'changes': form.changed_data,
            })
        )

    def delete_model(self, request, obj):
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            path=request.path,
            data=json.dumps({
                'model': obj._meta.model_name,
                'object_id': obj.pk,
            })
        )
        super().delete_model(request, obj)
def log_admin_action(request, obj, action, extra=None):
    payload = {
        'model': obj._meta.model_name,  # <-- đổi từ .label thành .model_name
        'object_id': obj.pk,
    }
    if extra:
        payload.update(extra)
    AuditLog.objects.create(
        user   = request.user,
        action = action,
        path   = request.path,
        data   = json.dumps(payload, ensure_ascii=False)
    )
class MealPriceAdmin( admin.ModelAdmin):
    list_display  = ('effective_date', 'daily_price', 'breakfast_price', 'lunch_price')
    list_editable = ('daily_price', 'breakfast_price', 'lunch_price')
    list_filter   = ('effective_date',)
    ordering      = ('-effective_date',)
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = 'change' if change else 'add'
        disp   = f"Cấu hình giá ăn - {obj.daily_price}"
        log_admin_action(request, obj, action, extra={
            'model':          'mealprice',
            'object_id':      obj.pk,
            'object_display': disp,
        })

    def delete_model(self, request, obj):
        disp = f"Cấu hình giá ăn - {obj.daily_price}"
        log_admin_action(request, obj, 'delete', extra={
            'model':          'mealprice',
            'object_id':      obj.pk,
            'object_display': disp,
        })
        super().delete_model(request, obj)
class StudentInline(admin.TabularInline):
    model = Student
    extra = 0                 # không sinh form trống
    fields = ('name',)        # chỉ hiện tên (có thể thêm các field khác)
    show_change_link = True   # có link vào form edit của từng student
class MyAdminSite(AdminSite):
    site_header = "Trang quản trị bữa ăn học sinh"
    index_title = "Bảng điều khiển"

    def index(self, request, extra_context=None):
        extra_context = extra_context or {}
        # Đổ nhanh 2 link vào dashboard
        extra_context['quick_links'] = [
            {
                'url': reverse('meals:statistics'),
                'label': '📊 Thống kê'
            },
            {
                'url': reverse('meals:student_payment_edit'),
                'label': '💳 Chỉnh sửa công nợ'
            }
        ]
        return super().index(request, extra_context)

    # Tuỳ ý bạn có thể override get_urls để thêm view custom,
    # nhưng ở đây chỉ cần index.

class ClassNameFilter(admin.SimpleListFilter):
    title = 'Lớp học'
    parameter_name = 'class_name'

    def lookups(self, request, model_admin):
        class_names = Student.objects.values_list('class_name', flat=True).distinct()
        return [(cls, cls) for cls in class_names if cls]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(class_name=self.value())
        return queryset
class StudentPaymentAdminForm(forms.ModelForm):
    class Meta:
        model = StudentPayment
        fields = '__all__'
        widgets = {
            # HTML5 month-picker
            'month': forms.TextInput(attrs={'type': 'month'}),
        }
        def validate_unique(self):
            # chỉ giữ lại validate field-level (nếu có), skip toàn bộ unique_together
            try:
                # gọi bản gốc, nhưng sẽ không raise vì ta ignore
                super().validate_unique()
            except forms.ValidationError:
                pass
from django.contrib.admin import SimpleListFilter

class StudentPaymentImportForm(forms.Form):
    month = forms.CharField(
        label="Tháng",
        widget=forms.TextInput(attrs={'type': 'month'})
    )
    classroom = forms.ModelChoiceField(
        queryset=ClassRoom.objects.all(),
        label="Lớp"
    )
    file = forms.FileField(label="File Excel")
class YearFilter(SimpleListFilter):
    title            = 'Năm'
    parameter_name   = 'year'
    template         = 'django_admin_listfilter_dropdown/dropdown_filter.html'
    def lookups(self, request, model_admin):
        # Lấy tập hợp các năm từ giá trị month “YYYY-MM”
        months = model_admin.model.objects.values_list('month', flat=True).distinct()
        years = sorted({m.split('-')[0] for m in months}, reverse=True)
        return [(y, y) for y in years]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(month__startswith=self.value() + '-')
        return queryset
class MonthFilter(SimpleListFilter):
    title            = 'Tháng'
    parameter_name   = 'month'
    template         = 'django_admin_listfilter_dropdown/dropdown_filter.html'
    def lookups(self, request, model_admin):
        year = request.GET.get('year')
        # chỉ lấy những month thuộc năm đã chọn (nếu có)
        months = model_admin.model.objects.values_list('month', flat=True).distinct()
        if year:
            months = [m for m in months if m.startswith(f"{year}-")]
        # chuyển thành list duy nhất, sort tăng dần
        unique = sorted({m for m in months})
        # trả về tuple (value, label) — ở đây label chỉ lấy phần MM
        return [(m, m.split('-')[1]) for m in unique]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(month=self.value())
        return queryset       
class StudentPaymentAdmin(admin.ModelAdmin):

    form = StudentPaymentAdminForm
    list_display  = (
        'student',
        'month',
        'tuition_fee_fmt',
        'meal_price',            # nếu đã override __str__ là "YYYY-MM-DD → X,000đ/ngày"
        'amount_paid_fmt',
        'remaining_balance_fmt',
    )
    search_fields = ('student__name','month')   # ← tìm theo tên học sinh hoặc tháng
    change_list_template = "admin/meals/studentpayment_changelist.html"
    list_filter = (
        YearFilter,
        ('student__classroom', RelatedDropdownFilter),
        MonthFilter,
    ) 
    verbose_name  = "Công nợ học sinh"
    verbose_name_plural = "Công nợ học sinh"
    @admin.display(ordering='tuition_fee', description='Học phí')
    def tuition_fee_fmt(self, obj):
        return f"{obj.tuition_fee:,.0f}"

    @admin.display(ordering='amount_paid', description='Số tiền đã đóng')
    def amount_paid_fmt(self, obj):
        return f"{obj.amount_paid:,.0f}"

    @admin.display(ordering='remaining_balance', description='Số dư')
    def remaining_balance_fmt(self, obj):
        return f"{obj.remaining_balance:,.0f}"
    def save_model(self, request, obj, form, change):
        existing = StudentPayment.objects.filter(
            student=obj.student,
            month=obj.month
        ).exclude(pk=obj.pk).first()

        if existing:
            # ghi đè lên existing
            existing.tuition_fee    = obj.tuition_fee
            existing.meal_price     = obj.meal_price
            existing.amount_paid    = obj.amount_paid
            existing.save()
            self.message_user(request, f"✅ Đã cập nhật bản ghi {existing.id}.", level=messages.SUCCESS)
            target = existing
        else:
            super().save_model(request, obj, form, change)
            target = obj

        # always log both add & change
        action = 'change' if change else 'add'
        log_admin_action(request, target, action, extra={
            'student_name':   target.student.name,
            'classroom_name': target.student.classroom.name,
            'tuition_fee':    str(target.tuition_fee),
            'amount_paid':    str(target.amount_paid),
            'meal_price':     str(target.meal_price.daily_price) if target.meal_price else '',
            'month':          str(target.month),
            'changed_fields': form.changed_data,
        })
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                'import-payments/',
                self.admin_site.admin_view(self.import_payments_view),
                name='meals_studentpayment_import'
            ),
        ]
        return custom + urls

    def import_payments_view(self, request):
        if request.method == 'POST':
            form = StudentPaymentImportForm(request.POST, request.FILES)
            if form.is_valid():
                month = form.cleaned_data['month']
                classroom = form.cleaned_data['classroom']
                wb = openpyxl.load_workbook(request.FILES['file'])
                sheet = wb.active

                overridden = []
                imported_count = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    student_name, raw_amount = row[:2]
                    if not student_name or raw_amount is None:
                        continue

                    # parse số tiền (loại bỏ dấu phẩy nếu có)
                    amount = Decimal(str(raw_amount).replace(",", ""))

                    try:
                        stu = Student.objects.get(classroom=classroom, name=student_name.strip())
                    except Student.DoesNotExist:
                        # bỏ qua những tên không khớp
                        continue

                    # kiểm tra overwrite
                    existed = StudentPayment.objects.filter(student=stu, month=month).exists()
                    if existed:
                        overridden.append(student_name)

                    # 1) Xác định meal_price_id  tuition_fee trước
                    prior = (StudentPayment.objects
                            .filter(student=stu)
                            .exclude(month=month)
                            .order_by('-month')
                            .first())
                    if prior:
                        price_id, fee = prior.meal_price_id, prior.tuition_fee
                    else:
                        price_id, fee = 2, 0

                    # 2) Tạo mới hoặc update, gán luôn cả price  fee trong defaults
                    sp, created = StudentPayment.objects.get_or_create(
                        student=stu,
                        month=month,
                        defaults={
                            'amount_paid':     amount,
                            'meal_price_id':   price_id,
                            'tuition_fee':     fee,
                        }
                    )
                    if not created:
                        sp.amount_paid     = amount
                        sp.meal_price_id   = price_id
                        sp.tuition_fee     = fee

                    # 3) Lưu, StudentPayment.save() của bạn sẽ tính remaining_balance mà không mắc lỗi
                    sp.save()
                    imported_count += 1

                # Recalc toàn bộ bảng (đảm bảo các tháng sau cũng update đúng)
                all_pays = StudentPayment.objects.all().order_by('student__id','month')
                for p in all_pays:
                    p.save()

                if overridden:
                    self.message_user(
                        request,
                        f"Sẽ override dữ liệu của: {', '.join(overridden)}",
                        level=messages.WARNING
                    )
                self.message_user(
                    request,
                    f"Import thành công {imported_count} bản ghi.",
                    level=messages.SUCCESS
                )
                return redirect('..')
        else:
            form = StudentPaymentImportForm()

        context = dict(
            self.admin_site.each_context(request),
            form=form,
            title="Import tiền đã đóng học sinh"
        )
        return TemplateResponse(request, "admin/meals/import_payments.html", context)
    def delete_model(self, request, obj):
        log_admin_action(request, obj, 'delete', extra={
            'student_name':   obj.student.name,
            'classroom_name': obj.student.classroom.name,
            'tuition_fee':    str(obj.tuition_fee),
            'amount_paid':    str(obj.amount_paid),
            'meal_price':     str(obj.meal_price.daily_price) if obj.meal_price else '',
            'month':          str(obj.month),
        })
        super().delete_model(request, obj)
class ClassRoomAdmin( admin.ModelAdmin):
    list_display  = ('name', 'term')
    list_filter   = (TermFilter,)
    search_fields = ('name', 'term',)
    ordering     = ('-id',)
    inlines = [StudentInline]  
    verbose_name = "Lớp học"
    verbose_name_plural = "Các lớp học"
    change_form_template = "admin/meals/classroom_change_form.html"
    def promote_students_choose_class_view(self, request, classroom_id):
        """
        Bước 2: Sau khi đã có danh sách student_ids trong session,
        hiển thị form chọn lớp đích. Khi POST, clone Student và clone StudentPayment +
        MealRecord tháng gần nhất, sau đó redirect về trang change của lớp gốc.
        """

        source_room = ClassRoom.objects.get(pk=classroom_id)
        student_ids = request.session.get('promote_student_ids', [])

        if request.method == 'POST':
            target_class_id = request.POST.get('target_class_id')
            if not target_class_id:
                messages.error(request, "Bạn phải chọn lớp đích.")
                return redirect(reverse('admin:meals_classroom_promote_choose_class',
                                        kwargs={'classroom_id': classroom_id}))
            try:
                target_room = ClassRoom.objects.get(pk=int(target_class_id))
            except (ValueError, ClassRoom.DoesNotExist):
                messages.error(request, "Lớp đích không hợp lệ.")
                return redirect(reverse('admin:meals_classroom_promote_choose_class',
                                        kwargs={'classroom_id': classroom_id}))

            # Lấy danh sách Student cần promote
            students_to_promote = Student.objects.filter(id__in=student_ids)
            new_students = []

            for old_student in students_to_promote:
                # 1) Clone Student (giữ các field cơ bản)
                new_student = Student.objects.create(
                    name=old_student.name,
                    classroom=target_room,
                    # Nếu Student có thêm field như birthday, address..., copy tương ứng
                    # birthday=old_student.birthday,
                    # address=old_student.address,
                    # ...
                )
                new_students.append(new_student)

                # 2) Lấy StudentPayment gần nhất của old_student
                latest_payment = (
                    StudentPayment.objects
                    .filter(student=old_student)
                    .order_by('-month')
                    .first()
                )

                if latest_payment:
                    # 3) Clone tạm StudentPayment ban đầu (giữ nguyên tất cả fields)
                    new_payment = StudentPayment.objects.create(
                        student=new_student,
                        month=latest_payment.month,
                        tuition_fee=latest_payment.tuition_fee,
                        amount_paid=latest_payment.amount_paid,
                        remaining_balance=latest_payment.remaining_balance,
                        meal_price=latest_payment.meal_price
                    )

                    # 4) Clone tất cả MealRecord của tháng gần nhất từ old_student sang new_student
                    year_str, month_str = latest_payment.month.split('-')
                    year_int = int(year_str)
                    month_int = int(month_str)

                    meal_records = MealRecord.objects.filter(
                        student=old_student,
                        date__year=year_int,
                        date__month=month_int
                    )
                    for mr in meal_records:
                        MealRecord.objects.create(
                            student=new_student,
                            date=mr.date,
                            meal_type=mr.meal_type,
                            status=mr.status,
                            non_eat=mr.non_eat,
                            absence_reason=mr.absence_reason
                        )

                    # ----- (5) VẶN RECALC CHO ĐÚNG: Ăn học = tuition_fee + spent_meals -----

                    # (i) Tính spent_meals_new (tổng tiền ăn của new_student trong tháng đó)
                    recs_new = MealRecord.objects.filter(
                        student=new_student,
                        date__year=year_int,
                        date__month=month_int,
                        meal_type__in=["Bữa sáng", "Bữa trưa"]
                    ).filter(
                        Q(status='Đủ') | Q(status='Thiếu', non_eat=2)
                    )
                    fee_b = Decimal(new_payment.meal_price.breakfast_price)
                    fee_l = Decimal(new_payment.meal_price.lunch_price)
                    spent_meals_new = sum(
                        fee_b if r.meal_type == "Bữa sáng" else fee_l
                        for r in recs_new
                    )

                    # (ii) Muốn giữ desired_remaining đúng bằng old_payment.remaining_balance
                    desired_remaining = latest_payment.remaining_balance

                    # (iii) Công thức recalc hệ thống: 
                    #      remaining_balance_new = amount_paid − (tuition_fee_new + spent_meals_new) + prior
                    #      với prior = 0 (thanh toán đầu tháng)
                    prior = Decimal('0')
                    new_tuition_fee = (
                        new_payment.amount_paid
                        - spent_meals_new
                        + prior
                        - desired_remaining
                    )

                    # (iv) Gán lại tuition_fee và remaining_balance cho new_payment
                    new_payment.tuition_fee = new_tuition_fee
                    new_payment.remaining_balance = desired_remaining
                    new_payment.save()
                    # ----- KẾT THÚC VẶN RECALC -----

            # Xóa session để tránh promote lại
            del request.session['promote_student_ids']

            messages.success(
                request,
                f"Đã tạo {len(new_students)} học sinh mới trong “{target_room.name}”, giữ nguyên dữ liệu cũ."
            )
            return redirect(reverse('admin:meals_classroom_change', args=(classroom_id,)))

        # Nếu GET: hiển thị form chọn lớp đích
        # 1️⃣ Lấy term từ GET hoặc mặc định = term của lớp nguồn
        selected_term = request.GET.get('term', source_room.term)

        # 2️⃣ Tất cả các term hiện có để đổ vào dropdown
        all_terms = (
            ClassRoom.objects
                    .values_list('term', flat=True)
                    .distinct()
                    .order_by('term')
        )

        # 3️⃣ Lọc ra chỉ các lớp cùng term, bỏ lớp nguồn
        target_rooms = (
            ClassRoom.objects
                    .filter(term=selected_term)
                    .exclude(pk=classroom_id)
                    .order_by('name')
        )

        context = {
            **self.admin_site.each_context(request),
            'opts':          self.model._meta,
            'original':      source_room,
            'all_terms':     all_terms,
            'selected_term': selected_term,
            'target_rooms':  target_rooms,
            'student_count': len(student_ids),
        }
        return render(request, "admin/meals/promote_students_choose_class.html", context)
    def promote_students_select_view(self, request, classroom_id):
        """
        Bước 1: Show một form list các Student của lớp <classroom_id> với checkbox.
        Khi submit (POST), redirect sang bước chọn lớp đích (choose_target_class).
        """
        room = ClassRoom.objects.get(pk=classroom_id)

        if request.method == 'POST':
            # Lấy danh sách student_id đã tick
            selected_ids = request.POST.getlist('students')
            if not selected_ids:
                messages.error(request, "Bạn phải chọn ít nhất một học sinh để lên lớp.")
                return redirect(f'../../{classroom_id}/promote_students/')

            # Chuyển sang bước 2, gắn student_ids vào session hay querystring
            # Ở đây mình sẽ dùng session để tạm giữ danh sách.
            request.session['promote_student_ids'] = selected_ids
            return redirect(reverse('myadmin:meals_classroom_promote_choose_class',
                                     kwargs={'classroom_id': classroom_id}))

        # GET: hiển thị form chứa danh sách student checkbox
        students_qs = Student.objects.filter(classroom=room).order_by('name')
        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'original': room,
            'students': students_qs,
        }
        # Template sẽ hiển thị checkbox list rồi POST lại về chính URL này
        return render(request, "admin/meals/promote_students_select.html", context)
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:classroom_id>/delete_students/',
                self.admin_site.admin_view(self.delete_students_view),
                name='meals_classroom_delete_students'
            ),
            path(
                '<int:classroom_id>/export_students/',
                self.admin_site.admin_view(self.export_students_view),
                name='meals_classroom_export_students'
            ),
            path(
                '<int:classroom_id>/import_students/',
                self.admin_site.admin_view(self.import_students_view),
                name='meals_classroom_import_students'
            ),
            path(
                '<int:classroom_id>/promote_students/',
                self.admin_site.admin_view(self.promote_students_select_view),
                name='meals_classroom_promote_students'
            ),
            path(
                '<int:classroom_id>/promote_students/choose_target_class/',
                self.admin_site.admin_view(self.promote_students_choose_class_view),
                name='meals_classroom_promote_choose_class'
                ),
        ]
        # Đặt custom URLs lên trước để không bị chặn bởi mặc định
        return custom + urls
    def delete_students_view(self, request, classroom_id):
        room = ClassRoom.objects.get(pk=classroom_id)
        if request.method == 'POST':
            # xóa tất cả sinh viên trong lớp, cascades luôn MealRecord & StudentPayment
            qs = Student.objects.filter(classroom=room)
            count, _ = qs.delete()
            self.message_user(request, f"Đã xóa {count} học sinh (và dữ liệu liên quan) của lớp {room.name}.")
            return HttpResponseRedirect(f'../../{classroom_id}/change/')

        context = {
            **self.admin_site.each_context(request),
            'opts':    self.model._meta,
            'original': room,
            'title':   f'Xác nhận xóa toàn bộ học sinh lớp “{room.name}”',
        }
        return TemplateResponse(request, "admin/meals/classroom_delete_students_confirmation.html", context)
    
    def export_students_view(self, request, classroom_id):
        room = ClassRoom.objects.get(pk=classroom_id)
        qs = Student.objects.filter(classroom=room).order_by('name')

        # Tạo workbook và sheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'Học sinh'

        # Header
        ws.append(['Tên học sinh'])

        # Dữ liệu
        for s in qs:
            ws.append([s.name])

        # Xuất file .xlsx
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"students_{room.name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = 'change' if change else 'add'
        disp   = obj.name
        log_admin_action(request, obj, action, extra={
            'model':          'classroom',
            'object_id':      obj.pk,
            'object_display': disp,
        })

    def delete_model(self, request, obj):
        disp = obj.name
        log_admin_action(request, obj, 'delete', extra={
            'model':          'classroom',
            'object_id':      obj.pk,
            'object_display': disp,
        })
        super().delete_model(request, obj)
    def import_students_view(self, request, classroom_id):
        room = ClassRoom.objects.get(pk=classroom_id)

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                self.message_user(request, "⚠️ Chưa chọn file Excel.", level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            # Đọc workbook
            wb = load_workbook(filename=excel_file, read_only=True, data_only=True)
            ws = wb.active

            count = 0
            # Bắt đầu từ row 2 để bỏ header
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[0]
                if name and str(name).strip():
                    Student.objects.get_or_create(name=str(name).strip(), classroom=room)
                    count += 1

            self.message_user(request, f"✅ Imported {count} học sinh vào lớp {room.name}.",
                               level=messages.SUCCESS)
            return HttpResponseRedirect(f'../../{classroom_id}/change/')

        # GET: render form
        context = {
            **self.admin_site.each_context(request),
            'opts':     self.model._meta,
            'original': room,
            'title':    f'Import học sinh cho lớp “{room.name}”',
        }
        return TemplateResponse(request, "admin/meals/import_students.html", context)

class StudentAdmin( admin.ModelAdmin):
    list_display = ('name','classroom')
    search_fields = ('name',)
    list_filter   = (
        ('classroom', RelatedDropdownFilter),
    )
    ordering        = ('-id',)
    # Đổi tên hiển thị của model Student trong Admin
    verbose_name = "Học sinh"
    verbose_name_plural = "Các học sinh"
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = 'change' if change else 'add'
        disp   = f"{obj.name} - {obj.classroom.name}"
        log_admin_action(request, obj, action, extra={
            'model':          'student',
            'object_id':      obj.pk,
            'object_display': disp,
        })

    def delete_model(self, request, obj):
        disp = f"{obj.name} - {obj.classroom.name}"
        log_admin_action(request, obj, 'delete', extra={
            'model':          'student',
            'object_id':      obj.pk,
            'object_display': disp,
        })
        super().delete_model(request, obj)

class MealRecordAdmin( admin.ModelAdmin):
    change_list_template = "admin/meals/mealrecord/change_list.html"
    list_display = ('student', 'date', 'meal_type', 'status')
    fields = ('student','date','meal_type','status','non_eat','absence_reason')
    list_filter  = (
        ('date', DropdownFilter),
        ('student__classroom', RelatedDropdownFilter),
    )
    ordering = ('-date',)
    date_hierarchy = 'date'
    search_fields = ('student__name',)
    # Đổi tên hiển thị của model MealRecord trong Admin
    verbose_name = "Bữa ăn"
    verbose_name_plural = "Các bữa ăn"
    def delete_model(self, request, obj):
        # Lấy loại bữa ăn (Sáng/Trưa)
        date_str = obj.date.strftime('%d/%m/%Y')
        mt       = obj.get_meal_type_display().replace('Bữa ', '').capitalize()
        disp = (
            f"Bản ghi bữa ăn - {date_str} - "
            f"{obj.student.name} - {obj.student.classroom.name} - {mt}"
        )
        log_admin_action(request, obj, 'delete', extra={
            'model':          'mealrecord',
            'object_id':      obj.pk,
            'object_display': disp,
        })
        super().delete_model(request, obj)
    def save_model(self, request, obj, form, change):
        existing_record = MealRecord.objects.filter(
             student=obj.student,
             date=obj.date,
             meal_type=obj.meal_type
        ).first()
        if existing_record:
            existing_record.status         = obj.status
            existing_record.non_eat        = obj.non_eat
            existing_record.absence_reason = obj.absence_reason
            existing_record.save()
            target = existing_record
        else:
            super().save_model(request, obj, form, change)
            target = obj

        # log hành động bữa ăn (add/change)
        action = 'change' if change else 'add'
        # Format ngày thành dd/mm/YYYY
        date_str = target.date.strftime('%d/%m/%Y')
        mt       = target.get_meal_type_display().replace('Bữa ', '').capitalize()
        disp = (
            f"Bản ghi bữa ăn - {date_str} - "
            f"{target.student.name} - {target.student.classroom.name} - {mt}"
        )
        log_admin_action(request, target, action, extra={
            'model':          'mealrecord',
            'object_id':      target.pk,
            'object_display': disp,
        })
        
        # Xác định chuỗi "YYYY-MM" từ obj.date
        year_month = obj.date.strftime("%Y-%m")
        from .models import StudentPayment
        try:
            sp = StudentPayment.objects.get(student=obj.student, month=year_month)
            sp.save()  # gọi model.save() sẽ dùng meal_price để tính remaining_balance
        except StudentPayment.DoesNotExist:
            pass
        
class MealRecordAdminForm(forms.ModelForm):
    class Meta:
        model = MealRecord
        fields = '__all__'
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 1) disable non_eat nếu status="Đủ"
        if self.instance and self.instance.status == "Đủ":
            self.fields['non_eat'].disabled = True
        # 2) ẩn luôn absence_reason nếu non_eat==0
        if self.instance and self.instance.non_eat == 0:
            self.fields.pop('absence_reason', None)
    
admin.site.register(ClassRoom, ClassRoomAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(MealRecord, MealRecordAdmin)
# Khởi tạo site mới
my_admin_site = MyAdminSite(name='myadmin')

# Đăng ký các model với site mới
from .admin import MealRecordAdmin  # form tuỳ chỉnh bạn đã có
my_admin_site.register(MealRecord, MealRecordAdmin)
try:
    my_admin_site.unregister(ClassRoom)
    my_admin_site.unregister(Student)
    my_admin_site.unregister(StudentPayment)
except:
    pass
@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ('timestamp', 'user', 'action_vi', 'object_display')
    search_fields   = ('data',)
    readonly_fields = ('timestamp', 'user', 'action', 'path', 'data')
    ordering = ('-timestamp',)
    def action_vi(self, obj):
        return {
            'add':    'Thêm',
            'change': 'Sửa',
            'delete': 'Xóa',
        }.get(obj.action, obj.action)
    action_vi.short_description = 'HÀNH ĐỘNG'
    
    @admin.display(description='ĐỐI TƯỢNG')
    def object_display(self, obj):
        info = json.loads(obj.data or '{}')
        # 1) Nếu view đã push sẵn object_display thì trả luôn
        if info.get('object_display'):
            return info['object_display']

        model = info.get('model')
        oid   = info.get('object_id')

        # 2) Student
        if model == 'student':
            stu = Student.objects.filter(pk=oid).first()
            return f"{stu.name} - {stu.classroom.name}" if stu else oid

        # 3) Classroom
        if model == 'classroom':
            cr = ClassRoom.objects.filter(pk=oid).first()
            return cr.name if cr else oid

        # 4) MealRecord
        if model == 'mealrecord':
            rec = MealRecord.objects.filter(pk=oid)\
                      .select_related('student','student__classroom')\
                      .first()
            if rec:
                mt = rec.get_meal_type_display().replace('Bữa ', '').capitalize()
                return f"Bản ghi bữa ăn - {rec.student.name} - {rec.student.classroom.name} - {mt}"
            return oid

        # 5) StudentPayment – parse và format số
        if model == 'studentpayment':
            s  = info.get('student_name','')
            c  = info.get('classroom_name','')
            t  = info.get('tuition_fee', '0')
            p  = info.get('amount_paid','0')
            m  = info.get('meal_price', '0')
            mo = info.get('month','')

            def fmt(x):
                try:
                    # nếu x là chuỗi số, convert rồi format dấu phẩy
                    return f"{float(x):,.2f}"
                except:
                    return x

            return (
                f"Thanh Toán - {s} - {c} - "
                f"{fmt(t)} - {fmt(p)} - {fmt(m)} - {mo}"
            )

        # 6) MealPrice
        if model == 'mealprice':
            price = MealPrice.objects.filter(pk=oid).first()
            return f"Cấu hình giá ăn - {price.daily_price}" if price else oid

        return oid
my_admin_site.register(StudentPayment, StudentPaymentAdmin)
# đăng ký ClassRoomAdmin lên my_admin_site
my_admin_site.register(ClassRoom, ClassRoomAdmin)
my_admin_site.register(Student, StudentAdmin)
# Đăng ký thêm StudentPayment nếu muốn
my_admin_site.register(User, UserAdmin)
my_admin_site.register(Group, GroupAdmin)
my_admin_site.register(MealPrice, MealPriceAdmin)
my_admin_site.register(AuditLog, AuditLogAdmin)