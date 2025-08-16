import openpyxl

try:
    # Load workbook
    wb = openpyxl.load_workbook('students_CAT HÈ 2025.xlsx')
    sheet = wb.active
    
    print(f'Sheet name: {sheet.title}')
    print(f'Max row: {sheet.max_row}')
    print(f'Max column: {sheet.max_column}')
    print('\nFirst 10 rows:')
    print('=' * 50)
    
    for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        print(f'Row {i}: {row}')
        
    print('\n' + '=' * 50)
    print('Checking for empty rows and data issues...')
    
    # Check for data issues
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not row[0] or row[0] is None:
            print(f'Row {i}: Empty student name')
        elif not row[1] or row[1] is None:
            print(f'Row {i}: Empty amount')
        else:
            print(f'Row {i}: Student="{row[0]}", Amount="{row[1]}" (Type: {type(row[1])})')
            
except Exception as e:
    print(f'Error: {e}')

